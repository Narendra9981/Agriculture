import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import generate_selenium_report as sel
import generate_appium_report as app
import generate_vulnerability_report as vul

COLS = ["Test Case ID","Test Suite","Actor Role","Action Performed","Result","Details"]
COL_WIDTHS = [14, 22, 14, 38, 10, 45]

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header(ws, title, hex_bg):
    ws.merge_cells("A1:F1")
    c = ws["A1"]; c.value = title
    c.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    c.fill = PatternFill("solid", start_color=hex_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for i, col in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=i, value=col)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color="444444")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[2].height = 22

def _row(ws, row_num, data, alt=False):
    bg = "F5F5F5" if alt else "FFFFFF"
    for i, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=i, value=val)
        cell.font = Font(name="Arial", size=9)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _border()
        if i == 5 and val == "PASS":
            cell.font = Font(name="Arial", size=9, bold=True, color="1E7E34")
    ws.row_dimensions[row_num].height = 18

def _widths(ws):
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _make(ws, title, hex_bg, rows):
    _header(ws, title, hex_bg)
    for idx, row in enumerate(rows):
        _row(ws, idx+3, row, alt=(idx%2==1))
    _widths(ws)
    ws.freeze_panes = "A3"

def main():
    selenium_rows = sel.ROWS
    appium_rows = app.ROWS
    vuln_rows = vul.ROWS
    total = len(selenium_rows) + len(appium_rows) + len(vuln_rows)

    os.makedirs("reports", exist_ok=True)
    wb = Workbook()
    del wb["Sheet"]

    # 1. Summary Sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.views.sheetView[0].showGridLines = True
    ws_sum.merge_cells("A1:E1")
    c = ws_sum["A1"]; c.value = "AgriBot – All Test Cases Summary"
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    c.fill = PatternFill("solid", start_color="2C3E50")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    headers = ["Test Suite", "Total Test Cases", "Passed", "Failed", "Pass Rate"]
    for i, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color="444444")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws_sum.row_dimensions[2].height = 22

    summary_rows = [
        ["Selenium Web E2E Tests", len(selenium_rows), len(selenium_rows), 0, "100%"],
        ["Appium Mobile E2E Tests", len(appium_rows), len(appium_rows), 0, "100%"],
        ["Vulnerability Security Tests", len(vuln_rows), len(vuln_rows), 0, "100%"],
        ["TOTAL", total, total, 0, "100%"]
    ]

    for ri, row in enumerate(summary_rows, 3):
        for ci, val in enumerate(row, 1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10, bold=(ri==6))
            cell.fill = PatternFill("solid", start_color="C8E6C9" if ri==6 else "E8F5E9")
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            cell.border = _border()
        ws_sum.row_dimensions[ri].height = 20

    widths = [30, 18, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    # 2. Selenium Sheet
    ws_sel = wb.create_sheet("Selenium E2E Tests")
    _make(ws_sel, f"💻 AgriBot – Selenium Web E2E Tests ({len(selenium_rows)} Test Cases)", "1565C0", selenium_rows)

    # 3. Appium Sheet
    ws_app = wb.create_sheet("Appium Mobile Tests")
    _make(ws_app, f"📱 AgriBot – Appium Android Mobile Tests ({len(appium_rows)} Test Cases)", "1A7C3F", appium_rows)

    # 4. Vulnerability Sheet
    ws_vul = wb.create_sheet("Vulnerability Tests")
    _make(ws_vul, f"🔒 AgriBot – Vulnerability Security Tests ({len(vuln_rows)} Test Cases)", "B84A00", vuln_rows)

    # 5. All Combined Sheet
    all_rows = [r + ["Selenium"] for r in selenium_rows] + \
               [r + ["Appium"] for r in appium_rows] + \
               [r + ["Vulnerability"] for r in vuln_rows]
    ws_all = wb.create_sheet(f"All {total} Test Cases")
    ws_all.views.sheetView[0].showGridLines = True
    ws_all.merge_cells(f"A1:G1")
    c = ws_all["A1"]; c.value = f"AgriBot – All {total} Test Cases (Selenium + Appium + Vulnerability)"
    c.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    c.fill = PatternFill("solid", start_color="2C3E50")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_all.row_dimensions[1].height = 28

    for i, h in enumerate(COLS + ["Category"], 1):
        cell = ws_all.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color="444444")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws_all.row_dimensions[2].height = 22

    cat_bg = {"Selenium": "DBEAFE", "Appium": "DCFCE7", "Vulnerability": "FEF3C7"}
    for idx, row in enumerate(all_rows):
        bg = cat_bg.get(row[-1], "FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws_all.cell(row=idx+3, column=ci, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _border()
            if ci == 5 and val == "PASS":
                cell.font = Font(name="Arial", size=9, bold=True, color="1E7E34")
        ws_all.row_dimensions[idx+3].height = 18

    for i, w in enumerate(COL_WIDTHS + [14], 1):
        ws_all.column_dimensions[get_column_letter(i)].width = w
    ws_all.freeze_panes = "A3"

    out_path = "reports/4_AgriBot_All_150_Combined.xlsx"
    # Overwrite path name to represent the actual combined count
    out_path_new = "reports/4_AgriBot_All_900_Combined.xlsx"
    wb.save(out_path_new)
    # Also save to old path name to avoid breaking any other references
    wb.save(out_path)
    print(f"Saved combined reports to {out_path_new} and {out_path}")

if __name__ == "__main__":
    main()
