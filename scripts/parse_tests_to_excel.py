import os
import glob
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def parse_xml_reports(directory):
    test_cases = []
    # Search for all TEST-*.xml files recursively under the directory
    search_path = os.path.join(directory, "**", "TEST-*.xml")
    xml_files = glob.glob(search_path, recursive=True)
    
    for xml_file in xml_files:
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            
            # The root is usually <testsuite>
            if root.tag == 'testsuite':
                suites = [root]
            else:
                # In case root is <testsuites> containing multiple <testsuite>
                suites = root.findall('testsuite')
                
            for suite in suites:
                suite_name = suite.attrib.get('name', 'UnknownSuite')
                for case in suite.findall('testcase'):
                    classname = case.attrib.get('classname', suite_name)
                    name = case.attrib.get('name', 'UnknownTest')
                    time_sec = float(case.attrib.get('time', 0.0))
                    
                    # Determine status
                    failure = case.find('failure')
                    error = case.find('error')
                    skipped = case.find('skipped')
                    
                    status = "PASSED"
                    err_msg = ""
                    stacktrace = ""
                    
                    if failure is not None:
                        status = "FAILED"
                        err_msg = failure.attrib.get('message', 'Failure')
                        stacktrace = failure.text or ""
                    elif error is not None:
                        status = "ERROR"
                        err_msg = error.attrib.get('message', 'Error')
                        stacktrace = error.text or ""
                    elif skipped is not None:
                        status = "SKIPPED"
                        err_msg = skipped.attrib.get('message', '')
                        
                    test_cases.append({
                        "Class Name": classname,
                        "Test Name": name,
                        "Status": status,
                        "Duration (s)": time_sec,
                        "Error Message": err_msg,
                        "Stacktrace": stacktrace
                    })
        except Exception as e:
            print(f"Error parsing {xml_file}: {e}")
            
    return test_cases

def write_sheet_data(ws, test_cases):
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = ["Class Name", "Test Name", "Status", "Duration (s)", "Error Message", "Stacktrace"]
    ws.append(headers)
    
    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if header in ["Status", "Duration (s)"] else left_align
        
    # Status styling
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
    pass_font = Font(name="Calibri", size=11, color="375623")
    
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red
    fail_font = Font(name="Calibri", size=11, color="C65911")
    
    skip_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Yellow
    skip_font = Font(name="Calibri", size=11, color="833C0C")
    
    for row_idx, case in enumerate(test_cases, 2):
        ws.cell(row=row_idx, column=1, value=case["Class Name"]).alignment = left_align
        ws.cell(row=row_idx, column=2, value=case["Test Name"]).alignment = left_align
        
        status_cell = ws.cell(row=row_idx, column=3, value=case["Status"])
        status_cell.alignment = center_align
        if case["Status"] == "PASSED":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif case["Status"] in ["FAILED", "ERROR"]:
            status_cell.fill = fail_fill
            status_cell.font = fail_font
        else:
            status_cell.fill = skip_fill
            status_cell.font = skip_font
            
        dur_cell = ws.cell(row=row_idx, column=4, value=case["Duration (s)"])
        dur_cell.alignment = center_align
        dur_cell.number_format = "0.000"
        
        ws.cell(row=row_idx, column=5, value=case["Error Message"]).alignment = left_align
        ws.cell(row=row_idx, column=6, value=case["Stacktrace"]).alignment = left_align
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            # If stacktrace or error msg is huge, cap width at 50 to keep sheet readable
            if cell.column in [5, 6]:
                val = val.split('\n')[0][:50]
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

def write_to_excel(test_cases, output_file):
    # Categorize test cases by type
    android_cases = []
    selenium_cases = []
    appium_cases = []
    
    for case in test_cases:
        cls = case["Class Name"]
        if cls.startswith("WebAuth") or cls.startswith("WebAdmin") or cls.startswith("WebCrop") or cls.startswith("WebForum") or cls.startswith("WebMarket") or "Selenium" in cls:
            selenium_cases.append(case)
        elif cls.startswith("MobileAuth") or cls.startswith("MobileCamera") or cls.startswith("MobileLocation") or cls.startswith("MobileOffline") or cls.startswith("MobileBot") or "Appium" in cls:
            appium_cases.append(case)
        else:
            android_cases.append(case)
            
    # 1. Write tabbed combined Excel report
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    if android_cases:
        ws_android = wb.create_sheet(title="Android Unit Tests")
        write_sheet_data(ws_android, android_cases)
    if selenium_cases:
        ws_selenium = wb.create_sheet(title="Selenium Web E2E")
        write_sheet_data(ws_selenium, selenium_cases)
    if appium_cases:
        ws_appium = wb.create_sheet(title="Appium Mobile E2E")
        write_sheet_data(ws_appium, appium_cases)
        
    if not wb.sheetnames:
        wb.create_sheet(title="No Results")
        
    wb.save(output_file)
    print(f"Saved combined tabbed test cases to {output_file}")
    
    # 2. Write separate Excel files
    out_dir = os.path.dirname(output_file) or "."
    
    # Selenium separate file
    wb_sel = openpyxl.Workbook()
    ws_sel = wb_sel.active
    ws_sel.title = "Selenium Web E2E"
    write_sheet_data(ws_sel, selenium_cases)
    sel_file = os.path.join(out_dir, "selenium-test-results.xlsx")
    wb_sel.save(sel_file)
    print(f"Saved {len(selenium_cases)} Selenium test cases to {sel_file}")
    
    # Appium separate file
    wb_app = openpyxl.Workbook()
    ws_app = wb_app.active
    ws_app.title = "Appium Mobile E2E"
    write_sheet_data(ws_app, appium_cases)
    app_file = os.path.join(out_dir, "appium-test-results.xlsx")
    wb_app.save(app_file)
    print(f"Saved {len(appium_cases)} Appium test cases to {app_file}")
    
    # Android separate file
    wb_and = openpyxl.Workbook()
    ws_and = wb_and.active
    ws_and.title = "Android Unit Tests"
    write_sheet_data(ws_and, android_cases)
    and_file = os.path.join(out_dir, "android-test-results.xlsx")
    wb_and.save(and_file)
    print(f"Saved {len(android_cases)} Android test cases to {and_file}")

if __name__ == "__main__":
    import sys
    search_dir = sys.argv[1] if len(sys.argv) > 1 else "app/build/test-results"
    out_file = sys.argv[2] if len(sys.argv) > 2 else "test-results.xlsx"
    
    cases = parse_xml_reports(search_dir)
    if not cases:
        print("No test results found to write!")
    else:
        write_to_excel(cases, out_file)
